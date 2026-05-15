import os
import io
import calendar
import requests
from datetime import date, datetime
from zoneinfo import ZoneInfo

import resend
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TZ = ZoneInfo("Europe/Bratislava")
REPORT_EMAIL = "prenako.kosice@gmail.com"

resend.api_key = os.environ.get("RESEND_API_KEY", "")
RESEND_FROM = "onboarding@resend.dev"


def _get_holidays(year: int) -> set:
    try:
        r = requests.get(
            f"https://date.nager.at/api/v3/PublicHolidays/{year}/SK", timeout=5
        )
        if r.status_code == 200:
            return {datetime.strptime(h["date"], "%Y-%m-%d").date() for h in r.json()}
    except Exception as e:
        print("Holiday API error:", e)
    return set()


def _classify(d: date, holidays: set) -> str:
    if d in holidays:
        return "Sviatok"
    if d.weekday() == 5:
        return "Sobota"
    if d.weekday() == 6:
        return "Nedeľa"
    return "Pracovný deň"


def send_open_shifts_alert(open_records: list):
    today = datetime.now(TZ).strftime("%d.%m.%Y")
    lines = [
        f"  - {r.employee.name} {r.employee.surname}: "
        f"začiatok {r.start_time.strftime('%H:%M')}, miesto {r.work_location}"
        for r in open_records
    ]
    body = (
        f"Nasledujúci zamestnanci sú stále prihlásení o 23:00 ({today}):\n\n"
        + "\n".join(lines)
        + "\n\nProsím, skontrolujte a manuálne ukončite smeny ak je to potrebné."
    )
    resend.Emails.send({
        "from": RESEND_FROM,
        "to": [REPORT_EMAIL],
        "subject": f"Otvorené smeny o 23:00 – {today}",
        "text": body,
    })


def _build_excel(year: int, month: int) -> bytes:
    from models import Employee, Attendance

    holidays = _get_holidays(year)
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])

    employees = Employee.query.order_by(Employee.surname, Employee.name).all()

    wb = Workbook()

    # Shared styles
    HDR_FILL = PatternFill("solid", fgColor="212529")
    FILLS = {
        "Pracovný deň": PatternFill("solid", fgColor="0D6EFD"),
        "Sobota":        PatternFill("solid", fgColor="FFD966"),
        "Nedeľa":        PatternFill("solid", fgColor="00FF15"),
        "Sviatok":       PatternFill("solid", fgColor="DC3545"),
    }
    DARK_TEXT = {"Sobota", "Nedeľa"}

    def _hdr_cell(ws, row, col, value):
        c = ws.cell(row, col, value)
        c.fill = HDR_FILL
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin()
        return c

    def _thin():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def _auto_width(ws):
        for col in ws.columns:
            width = max(len(str(c.value or "")) for c in col) + 4
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Prehľad"
    sum_hdrs = ["Zamestnanec", "Pracovné dni (h)", "Sobota (h)", "Nedeľa (h)", "Sviatky (h)", "Spolu (h)"]
    for col, h in enumerate(sum_hdrs, 1):
        _hdr_cell(ws_sum, 1, col, h)

    all_emp_data = []

    for emp in employees:
        records = (
            Attendance.query
            .filter(
                Attendance.employee_id == emp.id,
                Attendance.date >= first_day,
                Attendance.date <= last_day,
                Attendance.status == "completed",
            )
            .order_by(Attendance.date)
            .all()
        )

        normal_h = saturday_h = sunday_h = holiday_h = 0.0
        for rec in records:
            h = rec.hours_worked()
            t = _classify(rec.date, holidays)
            if t == "Pracovný deň":
                normal_h += h
            elif t == "Sobota":
                saturday_h += h
            elif t == "Nedeľa":
                sunday_h += h
            else:
                holiday_h += h

        total = round(normal_h + saturday_h + sunday_h + holiday_h, 2)
        all_emp_data.append((emp, records, normal_h, saturday_h, sunday_h, holiday_h, total))

        r = ws_sum.max_row + 1
        for col, val in enumerate([
            f"{emp.name} {emp.surname}",
            round(normal_h, 2), round(saturday_h, 2),
            round(sunday_h, 2), round(holiday_h, 2), total
        ], 1):
            c = ws_sum.cell(r, col, val)
            c.alignment = Alignment(horizontal="center")
            c.border = _thin()

    _auto_width(ws_sum)

    # ── Per-employee sheets ────────────────────────────────────────────────────
    detail_hdrs = ["Dátum", "Typ dňa", "Začiatok", "Koniec", "Hodiny", "Miesto práce"]

    for emp, records, *_ in all_emp_data:
        sheet_name = f"{emp.surname} {emp.name}"[:31]
        ws = wb.create_sheet(title=sheet_name)
        for col, h in enumerate(detail_hdrs, 1):
            _hdr_cell(ws, 1, col, h)

        for rec in records:
            day_type = _classify(rec.date, holidays)
            row_vals = [
                rec.date.strftime("%d.%m.%Y"),
                day_type,
                rec.start_time.strftime("%H:%M"),
                rec.end_time.strftime("%H:%M") if rec.end_time else "",
                rec.hours_worked(),
                rec.work_location,
            ]
            r = ws.max_row + 1
            fill = FILLS[day_type]
            colored_font = Font(color="000000" if day_type in DARK_TEXT else "FFFFFF")

            for col, val in enumerate(row_vals, 1):
                c = ws.cell(r, col, val)
                c.alignment = Alignment(horizontal="center")
                c.border = _thin()
                if col in (2, 5):
                    c.fill = fill
                    c.font = colored_font

        _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_monthly_report(year: int, month: int):
    excel_bytes = _build_excel(year, month)
    month_label = date(year, month, 1).strftime("%B %Y")
    filename = f"dochadzka_{year}_{month:02d}.xlsx"

    resend.Emails.send({
        "from": RESEND_FROM,
        "to": [REPORT_EMAIL],
        "subject": f"Mesačný výkaz dochádzky – {month_label}",
        "text": f"V prílohe nájdete výkaz dochádzky za {month_label}.",
        "attachments": [{"filename": filename, "content": list(excel_bytes)}],
    })